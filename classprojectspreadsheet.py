# Name: Elento Brent
# Date: June 9, 2026
# Program: Spreadsheet Automation Menu

from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference


# This function converts rainfall from inches to centimeters.
# Arguments:
#   data (float) - rainfall amount in inches
# Return Value:
#   float - rainfall amount in centimeters
def convertData(data):
    convertedValue = data * 2.54
    return convertedValue


# This function writes comma-separated data to a CSV file.
# Arguments:
#   filePath (str) - path to the CSV file
#   dataString (str) - comma-separated data to write
# Return Value:
#   bool - True if successful, False otherwise
def insertData(filePath, dataString):

    try:
        with open(filePath, "a") as csvFile:
            csvFile.write(dataString + "\n")
        return True

    except Exception as error:
        print("Error writing to file:", error)
        return False


# This function reads and displays the contents of a CSV file.
# Arguments:
#   filePath (str) - path to the CSV file
# Return Value:
#   None
def viewData(filePath):

    try:
        print("\nFile Path:", filePath)
        print("-" * 40)

        with open(filePath, "r") as csvFile:
            contents = csvFile.read()

            if contents:
                print(contents)
            else:
                print("The file is empty.")

    except Exception as error:
        print("Error reading file:", error)


# This function collects rainfall data from the user and stores it in a CSV file.
# Arguments:
#   None
# Return Value:
#   None
def getInput():

    filePath = "ZooData.csv"

    try:
        numberOfEntries = int(input("How many entries are being entered? "))

        for entry in range(numberOfEntries):

            entryDate = input("Enter the date: ")
            rainInches = float(input("Enter rainfall amount in inches: "))

            # Calling convertData(data) where data is the rainfall amount in inches.
            # Expected return value: rainfall converted from inches to centimeters.
            convertedRain = convertData(rainInches)

            csvData = f"{entryDate},{rainInches},{convertedRain:.2f}"

            insertSuccessful = insertData(filePath, csvData)

            if insertSuccessful:
                print(
                    f"\nThe following data was saved at "
                    f"{datetime.now()}: {csvData}\n"
                )

    except Exception as error:
        print("Error processing input:", error)


# This function creates an Excel chart from CSV data.
# Arguments:
#   filePath (str) - path to the CSV file
#   chartType (str) - chart type ('line' or 'bar')
# Return Value:
#   None
def createChart(filePath, chartType):

    dataChoice = input(
        "\nChoose data source:\n"
        "1. Rainfall Inches\n"
        "2. Rainfall Centimeters\n"
        "Enter choice: "
    )

    dates = []
    values = []

    try:

        with open(filePath, "r") as csvFile:

            for line in csvFile:

                rowData = line.strip().split(",")

                if len(rowData) >= 3:

                    dates.append(rowData[0])

                    if dataChoice == "1":
                        values.append(float(rowData[1]))
                        yAxisLabel = "Rainfall (Inches)"
                    else:
                        values.append(float(rowData[2]))
                        yAxisLabel = "Rainfall (Centimeters)"

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Rainfall Report"

        worksheet["A1"] = "Date"
        worksheet["B1"] = yAxisLabel

        for row in range(len(dates)):
            worksheet.cell(row=row + 2, column=1).value = dates[row]
            worksheet.cell(row=row + 2, column=2).value = values[row]

        data = Reference(
            worksheet,
            min_col=2,
            min_row=1,
            max_row=len(values) + 1
        )

        categories = Reference(
            worksheet,
            min_col=1,
            min_row=2,
            max_row=len(dates) + 1
        )

        if chartType == "bar":
            chart = BarChart()
        else:
            chart = LineChart()

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        chart.title = (
            f"elebre6186 "
            f"{datetime.now().strftime('%m/%d/%Y')}"
        )

        chart.x_axis.title = "Date"
        chart.y_axis.title = yAxisLabel

        worksheet.add_chart(chart, "D2")

        workbook.save("final.xlsx")

        print("\nChart created successfully!")
        print("Excel file saved as final.xlsx")

    except Exception as error:
        print("Error creating chart:", error)


# This function asks the user for a chart type and generates a report.
# Arguments:
#   filePath (str) - path to the CSV file
# Return Value:
#   None
def generateReport(filePath):

    chartChoice = input(
        "\nEnter chart type (line or bar): "
    ).lower()

    if chartChoice == "line":
        createChart(filePath, "line")

    elif chartChoice == "bar":
        createChart(filePath, "bar")

    else:
        print("Invalid chart type selected.")


def main():

    studentId = "elebre6186"

    print(f"{studentId} Spreadsheet Automation Menu.")
    print()

    menuOptions = [
        "Rainfall Spreadsheet",
        "View Stored Data",
        "Generate Report",
        "Save Spreadsheet",
        "Exit"
    ]

    for optionNumber, optionName in enumerate(menuOptions, start=1):
        print(f"{optionNumber}. {optionName}")

    print()

    # The next line retrieves the inputted option and stores into the variable called selectedOption.
    selectedOption = input("Please select an option (1-5): ")

    print()

    if selectedOption.isdigit():

        selectedOption = int(selectedOption)

        if 1 <= selectedOption <= len(menuOptions):

            print(f"You selected option number {selectedOption}.")
            print("The time and date is", str(datetime.now()))

            if selectedOption == 1:
                getInput()

            elif selectedOption == 2:
                viewData("ZooData.csv")

            elif selectedOption == 3:
                generateReport("ZooData.csv")

            else:
                print(
                    "Error: The chosen functionality "
                    "is not implemented yet"
                )

        else:
            print("Error: Invalid choice selected.")

    else:
        print("Error: Invalid choice selected.")


# Start the application
main()
